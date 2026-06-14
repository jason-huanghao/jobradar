"""Outputs router — Excel export and Markdown digest, scoped to a user's active profile."""

from __future__ import annotations

import io

from fastapi import APIRouter, Depends
from fastapi.responses import Response, StreamingResponse

from ...storage.db import get_session
from ...storage.repo import get_active_profile, list_scored
from ..deps import get_db_path, get_user_email

router = APIRouter()


@router.get("/excel")
def download_excel(
    db_path=Depends(get_db_path),
    user_email: str = Depends(get_user_email),
):
    """Export the user's scored jobs as an Excel file."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        return Response(content="openpyxl not installed", status_code=500)

    with next(get_session(db_path)) as session:
        profile = get_active_profile(session, user_email)
        rows = list_scored(session, profile.id) if profile else []

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Jobs"

    headers = [
        "Score", "Title", "Company", "Location", "Source", "Salary",
        "Skills", "Seniority", "Location Fit", "Language", "Visa", "Growth",
        "Status", "Reasoning", "URL",
    ]
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font

    for row_idx, (score, job) in enumerate(rows, 2):
        ws.append([
            round(score.overall, 1),
            job.title, job.company, job.location, job.source, job.salary,
            round(score.skills_match, 1), round(score.seniority_fit, 1),
            round(score.location_fit, 1), round(score.language_fit, 1),
            round(score.visa_friendly, 1), round(score.growth_potential, 1),
            score.status, score.reasoning, job.url,
        ])

    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=jobradar_export.xlsx"},
    )


@router.get("/digest")
def get_digest(
    min_score: float = 6.0,
    db_path=Depends(get_db_path),
    user_email: str = Depends(get_user_email),
):
    """Return a Markdown digest of the user's top-scoring jobs."""
    with next(get_session(db_path)) as session:
        profile = get_active_profile(session, user_email)
        rows = list_scored(session, profile.id, min_score=min_score)[:20] if profile else []

    lines = [f"# JobRadar Digest — Top {len(rows)} Matches\n"]
    for score, job in rows:
        lines.append(
            f"## {job.title} @ {job.company}  ·  ★ {score.overall:.1f}/10\n"
            f"📍 {job.location}  |  💰 {job.salary or 'N/A'}  |  🔗 [{job.url}]({job.url})\n\n"
            f"{score.reasoning}\n\n"
            f"**Apply angle:** {score.application_angle}\n\n---\n"
        )

    return Response(content="\n".join(lines), media_type="text/markdown")
